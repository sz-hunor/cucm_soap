from cryptography import x509
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import serialization
from datetime import datetime
from lxml import etree
from requests import Session
from requests.auth import HTTPBasicAuth
from requests.packages import urllib3
from zeep import Client
from zeep.cache import SqliteCache
from zeep.helpers import serialize_object
from zeep.transports import Transport
import base64
import getopt
import json
import openpyxl
import pprint
import re
import sys


def read_excel(file, sheet):
    """
    :param file: path to Excel file to be opened
    :param sheet: name of sheet to be read
    :return: list of dictionaries (JSON like object), one per row in the Excel sheet
    with the keys being the values of the first row and the values taken from each subsequent row
    has support for special syntax for nesting objects and declaring lists
    """
    output = []
    try:
        headers = [cell.value.split(":") for cell in openpyxl.load_workbook(file, data_only=True)[sheet][1]]
        row = [[cell.value for cell in row] for row in openpyxl.load_workbook(file, data_only=True)[sheet].iter_rows(min_row=2)]
    except (openpyxl.utils.exceptions.InvalidFileException, KeyError, FileNotFoundError) as err:
        print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: ERROR: {err}")
        sys.exit(2)

    for cell in row:
        dict_row = {}
        for key, value in zip(headers, cell):
            if value is None:
                continue
            else:
                # special cases
                value = None if str(value).lower() == "none" else value
                value = str(value) if not isinstance(value, str) else value
                value = [value] if re.search(r"^\[\'\[.*?\]\'\]$", str(key)) else value
                key = [str(element).strip("[]") for element in key]
                item = create_nested_dict(key, value)
                dict_row = merge_nested_dicts(dict_row, item)
        output.append(dict_row)
    return output


def create_nested_dict(key, value):
    """
    :param key: dictionary key
    :param value: dictionary value
    :return: a key: value pair broken up into a nested structure on ":" characters contained in the key
    """
    if not key:
        return value
    else:
        return {key[0]: create_nested_dict(key[1:], value)}


def merge_nested_dicts(dictionary_1, dictionary_2):
    """
    :param dictionary_1: dictionaries structured as JSON-like objects
    :param dictionary_2: dictionaries structured as JSON-like objects
    :return: a JSON-like dictionary resulting from the merger of the two input dictionaries
    a list is created out of the values of duplicate keys
    """
    merged_dict = dictionary_1.copy()
    for key, value in dictionary_2.items():
        if key in merged_dict:
            if isinstance(value, dict) and isinstance(merged_dict[key], dict):
                merged_dict[key] = merge_nested_dicts(merged_dict[key], value)
            elif isinstance(value, list) and isinstance(merged_dict[key], list):
                merged_dict[key].extend(value)
            else:
                if not isinstance(merged_dict[key], list):
                    merged_dict[key] = [merged_dict[key]]
                if not isinstance(value, list):
                    merged_dict[key].append(value)
                else:
                    merged_dict[key] = merged_dict[key] + value
        else:
            merged_dict[key] = value
    return merged_dict


def write_excel(dictionary, file, sheet, layers):
    """
    :param dictionary: dictionaries structured as JSON-like objects
    :param file: name of the file to write to
    :param sheet: name of the sheet to write to
    :param layers: number of layers in the dictionary encapsulating the useful data
    """

    for key, value in dictionary['return'].items():
        if isinstance(value, list):
            for item in value:
                create_excel(flatten_dict(item), file, sheet)
        elif isinstance(value, dict):
            create_excel(flatten_dict(value), file, sheet)
        elif isinstance(value, str) or value is None:
            print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: ERROR: Return: '{value}' does not contain enough data to add to output")

            
def remove_nesting(dictionary, layers):
    """
    :param dictionary: dictionaries structured as JSON-like objects
    :param layers: number of layers in the dictionary encapsulating the useful data
    :return: the JSON-like dictionary with the specified number of layers stripped
    """
    for i in range(layers):
        if dictionary is None or len(dictionary) == 0:
            return
        else:
            if isinstance(dictionary, list) or isinstance(dictionary, str):
                return dictionary
            dictionary = dictionary[next(iter(dictionary))]
    return dictionary


def flatten_dict(dictionary, parent_key=''):
    """
    :param dictionary: dictionaries structured as JSON-like objects
    :param parent_key: an operational argument needed for the recursion to work
    :return: a dictionary without any nested items and formatted with the special syntax for nesting objects and
    declaring lists
    """
    items = []
    for key, value in dictionary.items():
        new_key = f"{parent_key}:{key}" if parent_key else key
        if isinstance(value, dict):
            items.extend(flatten_dict(value, new_key).items())
        elif isinstance(value, list):
            for counter, item in enumerate(value):
                new_key = f"[{new_key}" if not re.search(r"^\[", new_key) else new_key
                if isinstance(item, dict):
                    items.extend(flatten_dict(item, f"{new_key}").items())
                else:
                    items.append((f"{new_key}]_{counter}", item))
        elif value is None:
            items.append((new_key, "none"))
        else:
            new_key = f"{new_key}]" if re.search(r"^\[", parent_key) else new_key
            items.append((new_key, value))
    return dict(items)


def create_excel(dictionary, file, sheet):
    """
    :param dictionary: flattened dictionary
    :param file: name of the file to write to
    :param sheet: name of the sheet to write to
    """
    # if default "Sheet" exists, delete it
    try:
        workbook = openpyxl.load_workbook(file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        del workbook["Sheet"]
    except openpyxl.utils.exceptions.InvalidFileException as err:
        print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: ERROR: {err}")
        sys.exit(2)

    # switch to sheet or create it if it doesn't exist
    if sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
    else:
        worksheet = workbook.create_sheet(sheet)

    # read in headers if they exist
    existing_headers = []
    for cell in worksheet[1]:
        if cell.value is None:
            continue
        else:
            existing_headers.append(cell.value)

    # append any new headers to existing headers
    new_headers = existing_headers.copy()
    for key in dictionary.keys():
        if re.search(r"_\d+$", key):
            key = re.sub(r"_\d+$", "", key)
        if key not in existing_headers:
            new_headers.append(key)

    # write header row to file
    for header_index, header in enumerate(new_headers):
        worksheet.cell(row=1, column=header_index + 1, value=header)

    # unpack the dictionary into a list of headers and values
    row_header = []
    row_value = []
    for key, value in dictionary.items():
        if re.search(r"_\d+$", key):
            key = re.sub(r"_\d+$", "", key)
        row_header.append(key)
        row_value.append(value)

    # find where to write row_value based on matching row_header to an item in new_headers and iterate rows
    next_row = worksheet.max_row + 1
    for header_index, header in enumerate(new_headers):
        try:
            value = row_value[row_header.index(header)]
        except ValueError:
            value = ""

        locate = header_index + 1
        while worksheet.cell(row=1, column=locate).value != header:
            locate += 1
        while worksheet.cell(row=next_row, column=locate).value:
            locate += 1
        worksheet.cell(row=next_row, column=locate, value=str(value))

    # save everything when done
    workbook.save(file)


def connect(cucm, username, password, verify, wsdl):
    """
    :param cucm: url or ip of the CUCM to connect to
    :param username: AXL enable username
    :param password: password for the AXL enabled username
    :param verify: weather the server certificate gets verified or not
    :param wsdl: location of the wsdl file
    :return: zeep connection object
    """
    if not verify:
        urllib3.disable_warnings()

    session = Session()
    session.verify = verify
    session.auth = HTTPBasicAuth(username, password)
    transport = Transport(session=session, timeout="none", cache=SqliteCache())
    client = Client(wsdl, transport=transport)
    service = client.create_service("{http://www.cisco.com/AXLAPIService/}AXLAPIBinding", cucm)
    return service


def check_if_element(request, xsd):
    """
    For reasons known only to the gods and cisco, if the soap function call contains certain elements, when sending
    the request, the request body needs to be sent as actual arguments instead of a JSON-like object
    This function checks in the xsd file if the submitted soap call contains any from a list of elements
    :param request: soap function call to be checked
    :param xsd: xml schema file to check against
    :return: True if any of the elements were found, otherwise False
    """

    request_xml_schema = get_xml_schema(request, xsd)

    if any(request_xml_schema.count(key_element) > 0 for key_element in ["choice", "searchCriteria", "SQL"]):
        return True
    else:
        return False


def get_xml_schema(request, xsd):
    """
    looks up the definition for the lxml.etree element in the provided xsd and returns it in string form
    :param request: soap function call to be checked
    :param xsd: xml schema file to do the lookup in
    :return: the definition of the xml element in text form
    """
    xsd_ns = {"xsd": "http://www.w3.org/2001/XMLSchema"}
    tree = etree.parse(xsd)

    element = tree.xpath(f'//xsd:element[@name="{request}"]', namespaces=xsd_ns)
    complex_type_ref = element[0].get("type").split(":")[1]
    complex_type = tree.xpath(f'//xsd:complexType[@name="{complex_type_ref}"]', namespaces=xsd_ns)

    decoded = etree.tostring(complex_type[0], pretty_print=True).decode("utf-8")
    return decoded


def soap_call(connection, payload, request, element):
    """
    Makes the SOAP request
    AXL/SOAP is a provisioning and configuration API, not a real-time API, beware that an attempt to make too many
    requests in quick succession could be throttled
    :param connection: zeep connection object
    :param payload: list of json-like nested dictionaries
    :param request: SOAP function call
    :param element: Boolean return of the check_if_element() function
    :return: possibly pain, possibly extra free time
    """
    result_list = []
    for row, item in enumerate(payload):
        try:
            if element:
                result = getattr(connection, request)(**item)
            else:
                result = getattr(connection, request)(item)

            print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: Information in row {row + 1} submitted")

            result = serialize_object(result, target_cls=dict)
            
            #The returned SOAP objects for any SQL querry have a different format to other objects, this section normalizes SQL soap objects
            if "SQL" in request:
                result = {'return':{str(i): {sql_item.tag: sql_item.text for sql_item in sql_list} for i, sql_list in enumerate(result['return']['row'])}}
            """
            Certificates stored in the SQL database are encoded in X.509 .der format, this section decodes the certificates
            The SQL database contains all information which is also in a decoded certificate except certificate validity periods, this section replaces
            the certificate field with validity information
            """
            if "certificate" in str(result):
                for sql_item in result['return'].values():
                    sql_item['certificate'] = sql_item['certificate'].replace("-----BEGIN CERTIFICATE-----", "").replace("-----END CERTIFICATE-----", "").strip()
                    cert = x509.load_der_x509_certificate(base64.b64decode(sql_item['certificate']), default_backend())
                    sql_item['certificate'] = {"Not Valid Before": cert.not_valid_before, "Not Valid After" : cert.not_valid_after}

            print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: Return: {pprint.pformat(result['return'], sort_dicts=False)}")
            result_list.append(result)
        except Exception as error:
            print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: ERROR: Error adding line: {str(error)}")
            result_list.append({"return": None})
    return result_list

def main(argv):
    """
    :param argv: short arguments must be specified in the format "-<argument> <value>", long in "--<argument>=<value>"
    """

    # default values can be specified in this section
    cucm = ""
    username = ""
    password = ""
    verify = False
    excel = ""
    sheet = ""
    wsdl = "AXLAPI.wsdl"
    xsd = "AXLSoap.xsd"
    request = ""
    preview = ""
    output = ""
    layers = 2
    req_json = ""

    try:
        opts, args = getopt.getopt(argv, "c:u:p:v:e:s:w:x:r:po:l:j:", ["cucm=", "user=", "pass=", "verify=", "excel=",
                                                                       "sheet=", "wsdl=", "xsd=", "request=", "preview",
                                                                       "output=", "remove_layers=", "req_json="])
    except getopt.GetoptError as err:
        print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: ERROR: {err}")
        sys.exit(2)

    for opt, arg in opts:
        if opt == "-c" or opt == "--cucm":
            cucm = "https://" + arg + ":8443/axl/"
        elif opt == "-u" or opt == "--user":
            username = arg
        elif opt == "-p" or opt == "--pass":
            password = arg
        elif opt == "-v" or opt == "--verify":
            if arg.lower() == "true":
                verify = True
            else:
                verify = arg
        elif opt == "-e" or opt == "--excel":
            excel = arg
        elif opt == "-s" or opt == "--sheet":
            sheet = arg
        elif opt == "-w" or opt == "--wsdl":
            wsdl = arg
        elif opt == "-x" or opt == "--xsd":
            xsd = arg
        elif opt == "-r" or opt == "--request":
            request = arg
        elif opt == "-p" or opt == "--preview":
            preview = True
        elif opt == "-o" or opt == "--output":
            output = arg
        elif opt == "-l" or opt == "--remove_layers":
            layers = arg
        elif opt == "-j" or opt == "--req_json":
            req_json = arg

    if preview:
        if all(excel and sheet):
            payload = read_excel(excel, sheet)
            print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: {json.dumps(payload, indent=4)}")
            return
        else:
            print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: ERROR: Preview mode: mandatory parameters missing")
            return

    if cucm and username and password and excel and sheet and wsdl and xsd and request:
        payload = read_excel(excel, sheet)
    elif cucm and username and password and req_json and wsdl and xsd and request:
        try:
            payload = [eval(req_json)]
        except SyntaxError as err:
            print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: ERROR: Error in the formatting of the JSON object: {err}")
            sys.exit(2)
    else:
        print(f"{datetime.now().strftime('%b %d %H:%M:%S')}: ERROR: Mandatory parameters missing")
        return

    connection = connect(cucm, username, password, verify, wsdl)
    check_for_element = check_if_element(request, xsd)
    result = soap_call(connection, payload, request, check_for_element)

    if output:
        for dictionary in result:
            write_excel(dictionary, file=output, sheet=request, layers=layers)


if __name__ == '__main__':
    main(sys.argv[1:])
